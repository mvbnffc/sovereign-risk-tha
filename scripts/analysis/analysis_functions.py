# Script with common analysis functions that will be used throughout.

import numpy as np
import rasterio
import os
import xarray as xr
from lmoments3 import distr
from scipy.stats import gumbel_r, kstest
import matplotlib.pyplot as plt
from copulas.bivariate import Clayton
import pandas as pd
import subprocess
from openpyxl import load_workbook
import datetime


def vectorized_damage(depth, value, heights, damage_percents):
    '''
    Vectorized damage function
    Apply damage function given a flood depth and exposure value.
    Function also needs as input the damage function heights > damage_percents
    '''
    # Use np.interp for vectorized linear interpolation
    damage_percentage = np.interp(depth, heights, damage_percents)
    return damage_percentage * value

def calculate_risk(flood, building_values, heights, damage_percents):
    '''
    Calculate flood risk given a flood map and a map of building values
    '''
    exposure = np.where(flood>0, building_values, 0)
    risk = vectorized_damage(flood, exposure, heights, damage_percents)

    return risk

def simple_risk_overlay(flood_path, exposure_path, output_path, damage_function, depth_adjuster):
    '''
    This function performs a simple risk overlay analysis.
    It takes as input a flood map, an exposure map, and a vulnerability curve.
    It outputs a risk raster
    '''
    # Load the rasters
    flood = rasterio.open(flood_path)
    exposure = rasterio.open(exposure_path)

    # Data info
    profile = flood.meta.copy()
    profile.update(dtype=rasterio.float32, compress='lzw', nodata=0)
    nodata = flood.nodata

    with rasterio.open(output_path, 'w', **profile) as dst:
        i = 0
        for ji, window in flood.block_windows(1):
            i += 1

            affine = rasterio.windows.transform(window, flood.transform)
            height, width = rasterio.windows.shape(window)
            bbox = rasterio.windows.bounds(window, flood.transform)

            profile.update({
                'height': height,
                'width': width,
                'affine': affine
            })

            flood_array = flood.read(1, window=window)
            exposure_array = exposure.read(1, window=window)
            flood_array = np.where(flood_array>0, flood_array, 0) # remove negative values
            flood_array = flood_array/depth_adjuster # convert to m
            risk = calculate_risk(flood_array, exposure_array, damage_function[0], damage_function[1]) # depths index 0 and prp damage index 1

            dst.write(risk.astype(rasterio.float32), window=window, indexes=1)

def calculate_total_raster_value(raster_path):
    '''
    This functions returns the total value of all non-zero raster cells in a layer.
    '''
    with rasterio.open(raster_path) as dst:
        raster = dst.read(1)
        raster_sum = np.sum(raster[raster>0])
    
    return raster_sum

def create_relative_exposure_layer(exposure_path, total_exposure_value, output_path):
    '''
    This function creates a relative exposure layer. For each cell it divides the value by the total summed values for the entire layer.
    '''
    # Load the raster
    exposure = rasterio.open(exposure_path)

    # Data info
    profile = exposure.meta.copy()
    profile.update(dtype=rasterio.float32, compress='lzw', nodata=0)
    nodata = exposure.nodata

    with rasterio.open(output_path, 'w', **profile) as dst:
        i = 0
        for ji, window in exposure.block_windows(1):
            i += 1

            affine = rasterio.windows.transform(window, exposure.transform)
            height, width = rasterio.windows.shape(window)
            bbox = rasterio.windows.bounds(window, exposure.transform)

            profile.update({
                'height': height,
                'width': width,
                'affine': affine
            })

            exposure_array = exposure.read(1, window=window)
            exposure_array = np.where(exposure_array>0, exposure_array/total_exposure_value, 0) # calculate proportional area rather than absolute (may have to adjust this for decimal precision)

            dst.write(exposure_array.astype(rasterio.float32), window=window, indexes=1)

def dignad_simple_risk_overlay(flood_path, exposure_path, output_path, damage_function, depth_adjuster, total_exposure_value):
    '''
    This function performs a simple risk overlay analysis. But instead of calculating the actual damage
    it calcualtes a proportional damage (related to total area or total distance) - this allows us to tweak the exposure information later in the analysis
    It takes as input a flood map, an exposure map, and a vulnerability curve.
    It outputs a risk raster
    '''
    # Load the rasters
    flood = rasterio.open(flood_path)
    exposure = rasterio.open(exposure_path)

    # Data info
    profile = flood.meta.copy()
    profile.update(dtype=rasterio.float32, compress='lzw', nodata=0)
    nodata = flood.nodata

    with rasterio.open(output_path, 'w', **profile) as dst:
        i = 0
        for ji, window in flood.block_windows(1):
            i += 1

            affine = rasterio.windows.transform(window, flood.transform)
            height, width = rasterio.windows.shape(window)
            bbox = rasterio.windows.bounds(window, flood.transform)

            profile.update({
                'height': height,
                'width': width,
                'affine': affine
            })

            flood_array = flood.read(1, window=window)
            exposure_array = exposure.read(1, window=window)
            exposure_array = np.where(exposure_array>0, exposure_array/total_exposure_value, 0) # calculate proportional area rather than absolute (may have to adjust this for decimal precision)
            flood_array = np.where(flood_array>0, flood_array, 0) # remove negative values
            flood_array = flood_array/depth_adjuster # convert to m
            risk = calculate_risk(flood_array, exposure_array, damage_function[0], damage_function[1]) # depths index 0 and prp damage index 1

            dst.write(risk.astype(rasterio.float32), window=window, indexes=1)

def simple_risk_overlay_percentile(flood_path, exposure_path, output_path, damage_function, depth_adjuster):
    '''
    TODO
    This function performs a simple risk overlay analysis, but calculates the 5th and 95th percentile of damages.
    It takes as input a flood map, an exposure map, and a vulnerability curve (with standard deviation information).
    It outputs two risk rasters - a 5th and 95th percentile
    '''
    return

def flopros_risk_overlay(flood_path, exposure_path, output_path, mask_path, damage_function, depth_adjuster):
    '''
    This function performs a risk overlay analysis. Before the risk analysis it masks all urban areas in the exposure dataset
    It takes as input a flood map, an exposure map, an urban area mask map, and a vulnerability curve.
    It outputs a risk raster
    '''
    # Load the rasters
    flood = rasterio.open(flood_path)
    exposure = rasterio.open(exposure_path)
    mask = rasterio.open(mask_path)

    # Data info
    profile = flood.meta.copy()
    profile.update(dtype=rasterio.float32, compress='lzw', nodata=0)
    nodata = flood.nodata

    with rasterio.open(output_path, 'w', **profile) as dst:
        i = 0
        for ji, window in flood.block_windows(1):
            i += 1

            affine = rasterio.windows.transform(window, flood.transform)
            height, width = rasterio.windows.shape(window)
            bbox = rasterio.windows.bounds(window, flood.transform)

            profile.update({
                'height': height,
                'width': width,
                'affine': affine
            })

            flood_array = flood.read(1, window=window)
            exposure_array = exposure.read(1, window=window)
            mask_array = mask.read(1, window=window)
            exposure_array = np.where(mask_array==1, 0, exposure_array) # wherever the urban mask equals 1, set to zero in exposure dataset
            flood_array = np.where(flood_array>0, flood_array, 0) # remove negative values
            flood_array = flood_array/depth_adjuster # convert to m 
            risk = calculate_risk(flood_array, exposure_array, damage_function[0], damage_function[1]) # depths index 0 and prp damage index 1

            dst.write(risk.astype(rasterio.float32), window=window, indexes=1)

def dryproofing_risk_overlay(flood_path, exposure_path, output_path, mask_path, damage_function, damage_function_dp, depth_adjuster):
    '''
    This function performs a risk overlay analysis. For masked dryproofed cells, the damage function is adjusted to return no damages for depths < 1m
    It takes as input a flood map, an exposure map, an urban area mask map, and two vulnerability curves.
    It outputs a risk raster
    '''
    # Load the rasters
    flood = rasterio.open(flood_path)
    exposure = rasterio.open(exposure_path)
    mask = rasterio.open(mask_path)

    # Data info
    profile = flood.meta.copy()
    profile.update(dtype=rasterio.float32, compress='lzw', nodata=0)
    nodata = flood.nodata

    with rasterio.open(output_path, 'w', **profile) as dst:
        i = 0
        for ji, window in flood.block_windows(1):
            i += 1

            affine = rasterio.windows.transform(window, flood.transform)
            height, width = rasterio.windows.shape(window)
            bbox = rasterio.windows.bounds(window, flood.transform)

            profile.update({
                'height': height,
                'width': width,
                'affine': affine
            })

            flood_array = flood.read(1, window=window)
            exposure_array = exposure.read(1, window=window)
            mask_array = mask.read(1, window=window)
            dry_proof_array = np.where(mask_array==1, exposure_array, 0) # cells for dry proofing are marked as 1
            non_dry_proof_array = np.where(mask_array==1, 0, exposure_array)
            flood_array = np.where(flood_array>0, flood_array, 0) # remove negative values
            flood_array = flood_array/depth_adjuster # convert to m
            dry_proof_risk = calculate_risk(flood_array, dry_proof_array, damage_function_dp[0], damage_function_dp[1]) # pass new damage function for dry proof cells
            non_dry_proof_risk = calculate_risk(flood_array, non_dry_proof_array, damage_function[0], damage_function[1]) # normal damage function
            # Sum risk arrays
            risk = dry_proof_risk + non_dry_proof_risk

            dst.write(risk.astype(rasterio.float32), window=window, indexes=1)

def relocation_risk_overlay(flood_path, exposure_path, output_path, mask_path, damage_function, depth_adjuster):
    '''
    This function performs a risk overlay analysis. Urban cells in the 2 year flood plain exposed to flood depths > 1 m are removed from the analysis.
    It takes as input a flood map, an exposure map, an urban cell (2yr flood > 1 m depth) map, and a vulnerability curves.
    It outputs a risk raster
    '''
    # Load the rasters
    flood = rasterio.open(flood_path)
    exposure = rasterio.open(exposure_path)
    mask = rasterio.open(mask_path)

    # Data info
    profile = flood.meta.copy()
    profile.update(dtype=rasterio.float32, compress='lzw', nodata=0)
    nodata = flood.nodata

    with rasterio.open(output_path, 'w', **profile) as dst:
        i = 0
        for ji, window in flood.block_windows(1):
            i += 1

            affine = rasterio.windows.transform(window, flood.transform)
            height, width = rasterio.windows.shape(window)
            bbox = rasterio.windows.bounds(window, flood.transform)

            profile.update({
                'height': height,
                'width': width,
                'affine': affine
            })

            flood_array = flood.read(1, window=window)
            exposure_array = exposure.read(1, window=window)
            mask_array = mask.read(1, window=window)
            exposure_array = np.where(mask_array>0, 0, exposure_array) # Remove cells that will be relocated
            flood_array = np.where(flood_array>0, flood_array, 0) # remove negative values
            flood_array = flood_array/depth_adjuster # convert to m
            risk = calculate_risk(flood_array, exposure_array, damage_function[0], damage_function[1])
            
            dst.write(risk.astype(rasterio.float32), window=window, indexes=1)

def combine_glofas(start, end, dir, area_filter):
    '''
    Function to combine glofas river discharge data into one xarray given a data directory with all the datasets in them
    as well as a start and end year for the desired discharge data. Also loads and clips the accumulating area dataset
    and masks the river discharge data by the specified upstream area threshold (area_filter)
    '''
    all_files = [os.path.join(dir, f"glofas_THA_{year}.grib") for year in range(start, end+1)] # if we do this for other countries will have to adjust filenames
    # Load all datasets into array
    datasets = [xr.open_dataset(file, engine='cfgrib') for file in all_files]
    # Concatenate all datasets along the time dimension
    combined_dataset = xr.concat(datasets, dim='time')
    # Make sure datasets are sorted by time
    combined_dataset = combined_dataset.sortby('time')
    # Load upstream area 
    upstream_area = xr.open_dataset(os.path.join(dir, "uparea_glofas_v4_0.nc"), engine='netcdf4') # might need to update the filename here
    # Get lat-lon limits from glofas data as will use this to clip the upstream area
    lat_limits = [combined_dataset.latitude.values[i] for i in [0, -1]]
    lon_limits = [combined_dataset.longitude.values[i] for i in [0, -1]]
    up_lats = upstream_area.latitude.values.tolist()
    up_lons = upstream_area.longitude.values.tolist()
    # Calculate slice indices
    lat_slice_index = [
    round((i-up_lats[0])/(up_lats[1]-up_lats[0]))
    for i in lat_limits
    ]
    lon_slice_index = [
        round((i-up_lons[0])/(up_lons[1]-up_lons[0]))
        for i in lon_limits
    ]
    # Slice upstream area to chosen glofas region
    red_upstream_area = upstream_area.isel(
        latitude=slice(lat_slice_index[0], lat_slice_index[1]+1),
        longitude=slice(lon_slice_index[0], lon_slice_index[1]+1),
    )
    # There are very minor rounding differences, so we update with the lat/lons from the glofas data
    red_upstream_area = red_upstream_area.assign_coords({
        'latitude': combined_dataset.latitude,
        'longitude': combined_dataset.longitude,
    })
    # Add the upstream area to the main data object and print the updated glofas data object:
    combined_dataset['uparea'] = red_upstream_area['uparea']
    # Mask the river discharge data
    combined_dataset_masked = combined_dataset.where(combined_dataset.uparea>=area_filter*1e6)


    return combined_dataset_masked

def check_timeseries(array, latitude, longitude):
    '''
    function for checking if the GloFAS timeseries is valid at a given lat-lon location
    '''
    test_point = array.sel(latitude=latitude, longitude=longitude, method='nearest')
    test_timeseries = test_point['dis24']
    test_acc = float(test_point['uparea'])
    # check for NaN values
    non_nan_count = test_timeseries.count().item()
    total_count = test_timeseries.size
    nan_ratio = non_nan_count/total_count

    # Does the timeseries pass the NaN threshold
    if nan_ratio < 1:
        return False, test_acc, "NaN values found"

    # Check for constant values
    if test_timeseries.min() == test_timeseries.max():
        return False, test_acc, "Constant timeseries values"

    # If all checks pass
    return True, test_acc, "Valid timeseries"

def perform_outlet_data_checks(outlets, discharge_data):
    '''
    function to perform the data checks on all basin outlet points. Returns list of basin outlets and whether it has a valid timeseries or not.
    '''
    # Loop through basins and check whether timeseries is valid
    results = []
    for index, row in outlets.iterrows():
        latitude = row['Latitude']
        longitude = row['Longitude']

        valid, acc, message = check_timeseries(discharge_data, latitude, longitude)

        # Store the results
        results.append({
            'HYBAS_ID': row['HYBAS_ID'],
            'Latitude': latitude,
            'Longitude': longitude,
            'Acc': acc,
            'Valid': valid,
            'Message': message
        })
        if not valid:
            print(f"ID: {row['HYBAS_ID']}, Lat: {latitude}, Lon: {longitude}, Acc: {acc}, Valid: {valid}, Message: {message}")
    
    return results

def extract_discharge_timeseries(outlets, discharge_data):
    '''
    function to extract discharge timeseries at basin outlet points. Returns a dictionary of timeseries with basin ID as key.
    '''

    # Dictionary to store timeseries data for each basin
    basin_timeseries = {}

    # Loop through basin outlets, storing each in turn
    for index, row in outlets.iterrows():
        basin_id = row['HYBAS_ID']
        lat = row['Latitude']
        lon = row['Longitude']
        point_data = discharge_data.sel(latitude=lat, longitude=lon, method='nearest')
        timeseries = point_data['dis24'].to_series()
        # store in dictionary
        basin_timeseries[basin_id] = timeseries
    
    return basin_timeseries

def fit_gumbel_distribution(basin_timeseries):
    '''
    Calculate extreme value distribution to all the basin timeseries. This function calculates the gumbel distribution and performs 
    the Kolomgorov-Smirnov test to check for the quality of fit. Returns a dictionary that reports each basin's gumbel parameters as 
    well as D and p-value from the Kolmogorov-Smirnov test. 
    '''
    # Initiate dictionaries
    gumbel_params = {}
    fit_quality = {}

    # Loop through basins, calculating annual maxima and fitting Gumbel distribution using L-moments
    for basin_id, timeseries in basin_timeseries.items():
        annual_maxima = timeseries.groupby(timeseries.index.year).max()

        # Fit Gumbel distribution using L-moments
        params = distr.gum.lmom_fit(annual_maxima)

        # Perform the Kolmogorov-Smirnov test (checking quality of fit)
        D, p_value = kstest(annual_maxima, 'gumbel_r', args=(params['loc'], params['scale']))

        gumbel_params[basin_id] = params
        fit_quality[basin_id] = (D, p_value)

    
    return gumbel_params, fit_quality

def calculate_uniform_marginals(basin_timeseries, gumbel_parameters):
    '''
    This function will transform annual maximum values from the discharge timeseries into uniform marginals
    for each river basin using the Cumulative Distribution Function of the fitted Gumbel distribution.
    '''
    # Initialize dictionary for uniform marginals
    uniform_marginals = {}

    for basin_id, timeseries in basin_timeseries.items():
        annual_maxima = timeseries.groupby(timeseries.index.year).max()
        params = gumbel_parameters[basin_id]
        uniform_marginals[basin_id] = gumbel_r.cdf(annual_maxima, loc=params['loc'], scale=params['scale'])

    return uniform_marginals

def calculate_basin_copula_pairs(uniform_marginals, plot_dependence_matrix=False):
    '''
    This function calculates the pairwise dependence of river basin discharge timeseries using the inverse Clayton Copula.
    The function takes as input the uniform marginals of all basins and outputs a dictionary of clayton copula models for each basin pair,
    a list of basins that throw an error and the clayton copula models stored as a matrix of basin combinations. If prompted, the function
    will also plot this dependence matrix visually.
    Following the approach in this paper -> https://onlinelibrary.wiley.com/doi/epdf/10.1111/risa.12382?saml_referrer
    '''
    # Initialize dictionary and list
    clayton_copula_models = {}
    clayton_error_basins = [] # list to store basins that cause an error
    
    for id1, margins1 in uniform_marginals.items():
        for id2, margins2 in uniform_marginals.items():
            if id1 < id2: # to avoid duplicate pairs
                try:
                    # Prepare the data for copula
                    data = np.column_stack((1-margins1, 1-margins2)) # interested in upper tail dependence so take inverse of CDF TODO: NEED TO CHECK IF THIS IS VALID
                    
                    # Fit the Clayton copula
                    flipped_clayton = Clayton()
                    flipped_clayton.fit(data)
        
                    # Store the copula model
                    clayton_copula_models[(id1, id2)] = flipped_clayton
                except ValueError as e:
                    clayton_error_basins.append((id1, id2))

    # Now going to store the copula pairs in a matrix
    basin_ids = list(uniform_marginals.keys()) # take the basin IDs from the uniform marginals dictionary
    N = len(basin_ids) # how many basin ids?

    # Initialize the matrix with NaNs
    dependence_matrix = np.full((N, N), np.nan)

    # Map from basin ID to matrix index
    id_to_index = {basin_id: index for index, basin_id in enumerate(basin_ids)}

    for (id1, id2), copula_model in clayton_copula_models.items():
        index1, index2 = id_to_index[id1], id_to_index[id2]
        dependence_matrix[index1, index2] = copula_model.theta
        dependence_matrix[index2, index1] = copula_model.theta

    # For error basins do the same but set theta to -1
    for (id1, id2) in clayton_error_basins:
        index1, index2 = id_to_index[id1], id_to_index[id2]
        dependence_matrix[index1, index2] = -1
        dependence_matrix[index2, index1] = -1

    # Debug (for infinity values) - not sure if needed but there are a few where I had to reassign basin outlets.
    dependence_matrix[np.isinf(dependence_matrix)] = 1000

    if plot_dependence_matrix:
        plt.imshow(dependence_matrix, cmap='viridis', interpolation='none', vmin=0, vmax=3)
        plt.colorbar()
        plt.title('Basin Dependence Structure')
        plt.xlabel('Basin Index')
        plt.ylabel('Basin Index')
        plt.show()

    return clayton_copula_models, clayton_error_basins, dependence_matrix

def minimax_ordering(dependence_matrix, basin_ids):
    '''
    This function creates the order of basins for aggregation in Monte Carlo simulation using the methodology described in Timonina et al (2015).
    This follows the following four steps:
    1. Choose the basin (i, j) pair with the maximum dependence out of all basin pairs.
    2. Choose the next basin (k) that is dependent on losses in basin i and j (but not basin i or j). We do this by pulling two vectors Vi and Vj which are the theta values 
    between all basins and basin i and j, respectively. We then construct a new vector that is the minimum of Vi and Vj. 
    3. We then maximize over this new vector to determine next basin
    4. Repeat iterations until we have done all basins

    The rationale behind this minimax approach is that we avoid understimating losses (but also don't unecessiraly overestimate by just ordering basins by dependence)
    '''

    # Step 1: Find the most dependent pair
    # Initialize a set to keep track of selected basin indices
    selected_indices = set()
    # convert dependence_matrix to a masked array, so that NaN values and -1 are not considered in the operation
    masked_dependence_matrix = np.ma.masked_less(dependence_matrix, 0) # masking out values < 0
    np.fill_diagonal(masked_dependence_matrix, np.ma.masked) # we want to ignore diagonal (NaN values)
    max_theta_index = np.unravel_index(np.argmax(masked_dependence_matrix, axis=None), masked_dependence_matrix.shape)
    ordered_basins = [basin_ids[max_theta_index[0]], basin_ids[max_theta_index[1]]]
    # Add indices to the set of selected indices
    selected_indices.update([max_theta_index[0], max_theta_index[1]])
    # Map from basin ID to matrix index
    id_to_index = {basin_id: index for index, basin_id in enumerate(basin_ids)}

    # Step 2-4: Loop until all basins are ordered
    while len(ordered_basins) < len(basin_ids):
        # Step 2: Choose basin k that is dependent on both basin i, j (last two basins in ordered_basins). Minimax approach
        # Exclude already selected basins from the selection process
        potential_next_indices = [i for i in range(len(basin_ids)) if i not in selected_indices]
        # Find the indices of the last two basins in ordered_basins
        last_two_indices = [id_to_index[basin] for basin in ordered_basins[-2:]]
        # Find dependency vectors for the last two basins
        dependency_vectors = masked_dependence_matrix[last_two_indices, :]
        # Calculate the minimum dependency for each row of the vector
        min_deps = np.ma.min(dependency_vectors, axis=0)
        # Mask already selected indices
        min_deps_masked = np.ma.copy(min_deps)
        # Debug 
        # Ensure min_deps_masked.mask is an array
        if np.isscalar(min_deps_masked.mask):
            min_deps_masked.mask = np.zeros(min_deps_masked.shape, dtype=bool)
        for idx in selected_indices:
            min_deps_masked.mask[idx] = True # mask the index if it's already in selected indices
        # Step 3: Find the maximum dependency value over the minimized vector - which will be the next basin
        next_basin_index = np.ma.argmax(min_deps_masked, fill_value=-np.inf)
        # Step 4: Continue iterations until there are no more basins left to process
        # Check if all options are effectively masked
        if min_deps_masked.mask.all():
            print("No suitable next basin found. Ending process.")
            break
        next_basin = basin_ids[next_basin_index]
        ordered_basins.append(next_basin)
        selected_indices.add(next_basin_index)
    
    return ordered_basins

def generate_conditional_sample(v, theta, r):
    '''
    Generate a conditional sample using the Flipped Calyton copula.
    Equation 12 from the Timonina et al (2015) paper

    :param v: Known loss in basin i
    :param theta: Copula parameter for dependency between basins i and j.
    :param r: Random value from uniform distribution for sampling.
    :retrun: Generated conditional loss in basin j.
    '''
    # Debug. Need to limit the size of theta to prevent overflow error
    if theta > 30:
        theta = 30 # need to determine what this threshold should be via trial and error.
    
    u = 1-(1+((1-v)**(-theta))*(((r**(-((theta)/(1+theta))))-1)))**(-(1/theta))

    return u 

def interpolate_damages(RPs, losses, sim_aep, protection_level=0.5):
    '''
    Function to interpolate damages between given an annual exceedance probability and 
    a depth-damage curve
    '''
    aeps = [1/i for i in RPs]
    # Ensure AEPs are in ascending order for np.interp
    aeps.sort() 
    losses = losses[::-1]

    # Interpolate based off simulated AEP
    if sim_aep >= protection_level: 
        return 0 
    else:
        interpolated_value = np.interp(sim_aep, aeps, losses)
        return interpolated_value
    
def d_interpolate_damages(RPs, losses, sim_aep, total_cs, proportional_cs, protection_level=0.5):
    '''
    Function to interpolate damages between given an annual exceedance probability and 
    a depth-damage curve. Adjusted for DIGNAD - also takes as input the capital stock value and proportional exposure info
    '''
    aeps = [1/i for i in RPs]
    # Ensure AEPs are in ascending order for np.interp
    aeps.sort() 
    losses = losses[::-1]

    # Multiply losses by capital stock
    losses = [i * (total_cs * proportional_cs) for i in losses]

    # Interpolate based off simulated AEP
    if sim_aep >= protection_level: 
        return 0 
    else:
        interpolated_value = np.interp(sim_aep, aeps, losses)
        return interpolated_value
    
def get_copula_model(copula_models, basin1, basin2):
    """
    Attempt to retrieve a copula model for a given pair of basins.
    Tries both possible orders of the basin IDs.

    :param copula_models: Dictionary of copula models.
    :param basin1: ID of the first basin.
    :param basin2: ID of the second basin.
    :return: The copula model if found, else None.
    """
    return copula_models.get((basin1, basin2)) or copula_models.get((basin2, basin1))

def basin_loss_curve(loss_df, basin_id, basin_col, epoch_val, scenario_val, urban_class, rps):
    '''
    Function for extracting loss curves from each basin from the risk results dataframe.
    Extracting these loss curves at the beginning of the Monte Carlo simulation significantly reduces 
    compuatation times. 
    '''
    losses = {} # initialize empty dictionary to store losses and protection level
    basin_df = loss_df[(loss_df[basin_col]==basin_id) & (loss_df['epoch']==epoch_val) & (loss_df['adaptation_scenario']==scenario_val) & (loss_df['urban_class']==urban_class)]
    grouped_basin_df = basin_df.groupby([basin_col, 'RP', 'Pr_L_AEP']).agg({'damages':'sum'}).reset_index()
    # # Pull unique protection levels from the grouped dataframe
    unique_protection_levels = grouped_basin_df['Pr_L_AEP'].unique()
    if len(unique_protection_levels) == 0:
        unique_protection_levels = [1]
    for i in unique_protection_levels:
        losses[i] = [grouped_basin_df.loc[(grouped_basin_df['RP'] == rp) & (grouped_basin_df['Pr_L_AEP']==i), 'damages'].sum() for rp in rps]
    
    return losses

def monte_carlo_dependence_simulation(loss_df, rps, basin_col, epoch_val, scenario_val, urban_class, protection_level, num_years, ordered_basins, copula_models, num_simulations=10000):
    '''
    Perform Monte Carlo simulations of yearly losses incorporating basin dependencies.

    :param loss_df: dataframe with losses from risk analysis
    :param rps: list of return periods to consider. 
    :param basin_col: name of column for basins (e.g. 'HB_L6')
    :param epoch_val: name of epoch value (e.g. 'Today')
    :param scenario_val: name of scenario (e.g. 'Baseline')
    :param urban_class: name of urban class to consider (e.g. 'Residential')
    :param protection_level: what is the baseline protection level (e.g. 0.5 or 1 in 2 years)
    :param num_years: Number of years to simulate
    :param ordered_basins: List of basin IDs ordered by dependency
    :param copula_models: Dictionary holding copula model for each basin pair
    :param num_simulations: Number of simulations (default is 10,000).
    :return: Dataframe of simulated national losses for each year.
    '''

    # To speed up the Monte-Carlo simulation we are going to pre-compute some variables
    # precompute loss-probability curves for each basin
    basin_loss_curves = {basin_id: basin_loss_curve(loss_df, basin_id, basin_col, epoch_val, scenario_val, urban_class, rps) for basin_id in ordered_basins}
    # Initialize array for national losses
    national_losses_per_year = np.zeros((num_simulations, num_years))
    # Generate all random numbers in advance
    random_numbers = np.random.uniform(0, 1, (num_simulations, num_years, len(ordered_basins))).astype(np.float32)

    for simulation in range(num_simulations):
        # # print progress
        # if simulation % 50 == 0:
        #     print('Simulation progress: %s out of %s' % (simulation, num_simulations))
        for year in range(num_years):
            # Initialize a list to store losses for each basin for the current year
            yearly_losses = []
            yearly_loss_values = []
            for i, basin_id in enumerate(ordered_basins):
                # print(basin_id)
                if i == 0:
                    # Handle first basin
                    r = random_numbers[simulation, year, i]
                    loss_curves = basin_loss_curves[basin_id]
                    basin_loss = 0
                    yearly_losses.append(r) # add current loss simulation to the list
                    for Pr_L in loss_curves: # loop through basin protection levels
                        if Pr_L <= r:
                            # print(Pr_L, 'smaller than', r, 'continuing...') # if baseline protection is achieved...
                            continue
                        else:
                            yearly_loss_values.append(interpolate_damages(rps, loss_curves[Pr_L], r, protection_level))
                            
                else:
                    loss_curves = basin_loss_curves[basin_id]
                    # Handle subsequent basins with dependencies
                    copula = get_copula_model(copula_models, ordered_basins[i-1], basin_id)
                    if copula is not None:
                        # Apply dependency model if theta exists
                        r = random_numbers[simulation, year, i]
                        previous_loss = yearly_losses[i-1]
                        current_loss = generate_conditional_sample(previous_loss, copula.theta, r)
                        yearly_losses.append(current_loss)
                        # TODO: need to check below assumption. Currently, the (1-current_loss) criteria leads to stupid results.
                        # in the below interpolation the (1-current_loss) part of the equation is critical.
                        # because the copula is optimized to model tail dependencies (e.g. > 0.9) and our AEPs are 
                        # essentially inverted (e.g. 0.001 is extreme) we need to invert the random number for interpolating the
                        # losses. This changes nothing apart from ensuring tail dependency is preserved. 
                        for Pr_L in loss_curves: # loop through basin protection levels
                            if Pr_L <= current_loss:
                                # print(Pr_L, 'smaller than', r, 'continuing...') # if baseline protection is achieved...
                                continue
                            else:
                                yearly_loss_values.append(interpolate_damages(rps, loss_curves[Pr_L], current_loss, protection_level))
                    else:
                        # Independent simulation for this basin
                        r = random_numbers[simulation, year, i]
                        yearly_losses.append(r)
                        for Pr_L in loss_curves: # loop through basin protection levels
                            if Pr_L <= r:
                                continue
                            else:
                                yearly_loss_values.append(interpolate_damages(rps, loss_curves[Pr_L], r, protection_level))

            # Aggregate losses for the current year
            national_losses_per_year[simulation, year] = sum(yearly_loss_values)

    # Convert the results into a DataFrame
    return pd.DataFrame(national_losses_per_year, columns=[f'Year_{i+1}' for i in range(num_years)])

def urban_monte_carlo_dependence_simulation(loss_df, rps, basin_col, epoch_val, scenario_val, urban_class, protection_level, num_years, ordered_basins, copula_models, num_simulations=10000):
    '''
    Adjusted to account for urban protection
    Perform Monte Carlo simulations of yearly losses incorporating basin dependencies. This function is specifically for simulating urban flood protection

    :param loss_df: dataframe with losses from risk analysis
    :param rps: list of return periods to consider. 
    :param basin_col: name of column for basins (e.g. 'HB_L6')
    :param epoch_val: name of epoch value (e.g. 'Today')
    :param scenario_val: name of scenario (e.g. 'Baseline')
    :param urban_class: name of urban class to consider (e.g. 'Residential')
    :param protection_level: what is the baseline protection level (e.g. 0.5 or 1 in 2 years)
    :param num_years: Number of years to simulate
    :param ordered_basins: List of basin IDs ordered by dependency
    :param copula_models: Dictionary holding copula model for each basin pair
    :param num__simulations: Number of simulations (default is 10,000).
    :return: Dataframe of simulated national losses for each year.
    '''

    # To speed up the Monte-Carlo simulation we are going to pre-compute some variables
    # precompute loss-probability curves for each basin
    urban_basin_loss_curves = {basin_id: basin_loss_curve(loss_df, basin_id, basin_col, epoch_val, scenario_val, urban_class, rps) for basin_id in ordered_basins}
    basin_loss_curves = {basin_id: basin_loss_curve(loss_df, basin_id, basin_col, epoch_val, 'Baseline', urban_class, rps) for basin_id in ordered_basins}
    # Initialize array for national losses
    national_losses_per_year = np.zeros((num_simulations, num_years))
    # Generate all random numbers in advance
    random_numbers = np.random.uniform(0, 1, (num_simulations, num_years, len(ordered_basins))).astype(np.float32)

    for simulation in range(num_simulations):
        # # print progress
        # if simulation % 50 == 0:
        #     print('Simulation progress: %s out of %s' % (simulation, num_simulations))
        for year in range(num_years):
            # Initialize a list to store losses for each basin for the current year
            yearly_losses = []
            yearly_loss_values = []
            for i, basin_id in enumerate(ordered_basins):
                if i == 0:
                    # Handle first basin
                    r = random_numbers[simulation, year, i]
                    urban_loss_curves = urban_basin_loss_curves[basin_id]
                    loss_curves = basin_loss_curves[basin_id]
                    yearly_losses.append(r) # add current loss simulation to the list
                    if r < 0.01: # use baseline maps if AEP < 0.01
                        for Pr_L in loss_curves:
                            yearly_loss_values.append(interpolate_damages(rps, loss_curves[Pr_L], r, protection_level))
                    else:
                        for Pr_L in urban_loss_curves:
                            if Pr_L <= r:
                                continue
                            else:
                                yearly_loss_values.append(interpolate_damages(rps, urban_loss_curves[Pr_L], r, protection_level))
                            
                else:
                    urban_loss_curves = urban_basin_loss_curves[basin_id]
                    loss_curves = basin_loss_curves[basin_id]
                    # Handle subsequent basins with dependencies
                    copula = get_copula_model(copula_models, ordered_basins[i-1], basin_id)
                    if copula is not None:
                        # Apply dependency model if theta exists
                        r = random_numbers[simulation, year, i]
                        previous_loss = yearly_losses[i-1]
                        current_loss = generate_conditional_sample(previous_loss, copula.theta, r)
                        yearly_losses.append(current_loss)
                        # TODO: need to check below assumption. Currently, the (1-current_loss) criteria leads to stupid results.
                        # in the below interpolation the (1-current_loss) part of the equation is critical.
                        # because the copula is optimized to model tail dependencies (e.g. > 0.9) and our AEPs are 
                        # essentially inverted (e.g. 0.001 is extreme) we need to invert the random number for interpolating the
                        # losses. This changes nothing apart from ensuring tail dependency is preserved. 
                        if current_loss < 0.01: # Use baseline maps
                            for Pr_L in loss_curves: 
                                yearly_loss_values.append(interpolate_damages(rps, loss_curves[Pr_L], current_loss, protection_level))
                        else:
                            for Pr_L in urban_loss_curves:
                                if Pr_L <= current_loss:
                                    continue
                                else:
                                    yearly_loss_values.append(interpolate_damages(rps, urban_loss_curves[Pr_L], current_loss, protection_level))                    
                    else:
                        # Independent simulation for this basin
                        r = random_numbers[simulation, year, i]
                        yearly_losses.append(r)
                        if r < 0.01: # use baseline maps if AEP < 0.01
                            for Pr_L in loss_curves:
                                yearly_loss_values.append(interpolate_damages(rps, loss_curves[Pr_L], r, protection_level))
                        else:
                            for Pr_L in urban_loss_curves:
                                if Pr_L <= r:
                                    continue
                                else:
                                    yearly_loss_values.append(interpolate_damages(rps, urban_loss_curves[Pr_L], r, protection_level))

            # Aggregate losses for the current year
            national_losses_per_year[simulation, year] = sum(yearly_loss_values)

    # Convert the results into a DataFrame
    return pd.DataFrame(national_losses_per_year, columns=[f'Year_{i+1}' for i in range(num_years)])

def sectoral_monte_carlo_dependence_simulation(loss_df, rps, basin_col, epoch_val, scenario_val, protection_level, num_years, ordered_basins, copula_models, num_simulations=10000):
    '''
    Perform Monte Carlo simulations of yearly losses incorporating basin dependencies.
    Same as the monte carlo dependence simulation, except it now outputs sectoral losses.

    :param loss_df: dataframe with losses from risk analysis
    :param rps: list of return periods to consider. 
    :param basin_col: name of column for basins (e.g. 'HB_L6')
    :param epoch_val: name of epoch value (e.g. 'Today')
    :param scenario_val: name of scenario (e.g. 'Baseline')
    :param protection_level: what is the baseline protection level (e.g. 0.5 or 1 in 2 years)
    :param num_years: Number of years to simulate
    :param ordered_basins: List of basin IDs ordered by dependency
    :param copula_models: Dictionary holding copula model for each basin pair
    :param num_simulations: Number of simulations (default is 10,000).
    :return: Dataframe of simulated national losses for each year, for each sector
    '''

    # To speed up the Monte-Carlo simulation we are going to pre-compute some variables
    # precompute loss-probability curves for each basin for the four ectors
    
    res_basin_loss_curves = {basin_id: basin_loss_curve(loss_df, basin_id, basin_col, epoch_val, scenario_val, "Residential", rps) for basin_id in ordered_basins}
    com_basin_loss_curves = {basin_id: basin_loss_curve(loss_df, basin_id, basin_col, epoch_val, scenario_val, "Commercial", rps) for basin_id in ordered_basins}
    ind_basin_loss_curves = {basin_id: basin_loss_curve(loss_df, basin_id, basin_col, epoch_val, scenario_val, "Industrial", rps) for basin_id in ordered_basins}
    inf_basin_loss_curves = {basin_id: basin_loss_curve(loss_df, basin_id, basin_col, epoch_val, scenario_val, "Infrastructure", rps) for basin_id in ordered_basins}
    
    # Initialize arrays for national losses
    res_national_losses_per_year = np.zeros((num_simulations, num_years))
    com_national_losses_per_year = np.zeros((num_simulations, num_years))
    ind_national_losses_per_year = np.zeros((num_simulations, num_years))
    inf_national_losses_per_year = np.zeros((num_simulations, num_years))
    # Generate all random numbers in advance
    random_numbers = np.random.uniform(0, 1, (num_simulations, num_years, len(ordered_basins))).astype(np.float32)

    for simulation in range(num_simulations):
        # # print progress
        # if simulation % 50 == 0:
        #     print('Simulation progress: %s out of %s' % (simulation, num_simulations))
        for year in range(num_years):
            # Initialize a list to store losses for each basin for the current year
            yearly_losses = []
            res_yearly_loss_values = []
            com_yearly_loss_values = []
            ind_yearly_loss_values = []
            inf_yearly_loss_values = []
            for i, basin_id in enumerate(ordered_basins):
                # print(basin_id)
                if i == 0:
                    # Handle first basin
                    r = random_numbers[simulation, year, i]
                    res_loss_curves = res_basin_loss_curves[basin_id]
                    com_loss_curves = com_basin_loss_curves[basin_id]
                    ind_loss_curves = ind_basin_loss_curves[basin_id]
                    inf_loss_curves = inf_basin_loss_curves[basin_id]
                    basin_loss = 0
                    yearly_losses.append(r) # add current loss simulation to the list
                    for Pr_L in res_loss_curves: # loop through basin protection levels
                        if Pr_L <= r:
                            # print(Pr_L, 'smaller than', r, 'continuing...') # if baseline protection is achieved...
                            continue
                        else:
                            res_yearly_loss_values.append(interpolate_damages(rps, res_loss_curves[Pr_L], r, protection_level))
                            com_yearly_loss_values.append(interpolate_damages(rps, com_loss_curves[Pr_L], r, protection_level))
                            ind_yearly_loss_values.append(interpolate_damages(rps, ind_loss_curves[Pr_L], r, protection_level))
                            inf_yearly_loss_values.append(interpolate_damages(rps, inf_loss_curves[Pr_L], r, protection_level))
                            
                else:
                    res_loss_curves = res_basin_loss_curves[basin_id]
                    com_loss_curves = com_basin_loss_curves[basin_id]
                    ind_loss_curves = ind_basin_loss_curves[basin_id]
                    inf_loss_curves = inf_basin_loss_curves[basin_id]
                    # Handle subsequent basins with dependencies
                    copula = get_copula_model(copula_models, ordered_basins[i-1], basin_id)
                    if copula is not None:
                        # Apply dependency model if theta exists
                        r = random_numbers[simulation, year, i]
                        previous_loss = yearly_losses[i-1]
                        current_loss = generate_conditional_sample(previous_loss, copula.theta, r)
                        yearly_losses.append(current_loss)
                        # TODO: need to check below assumption. Currently, the (1-current_loss) criteria leads to stupid results.
                        # in the below interpolation the (1-current_loss) part of the equation is critical.
                        # because the copula is optimized to model tail dependencies (e.g. > 0.9) and our AEPs are 
                        # essentially inverted (e.g. 0.001 is extreme) we need to invert the random number for interpolating the
                        # losses. This changes nothing apart from ensuring tail dependency is preserved. 
                        for Pr_L in res_loss_curves: # loop through basin protection levels
                            if Pr_L <= current_loss:
                                # print(Pr_L, 'smaller than', r, 'continuing...') # if baseline protection is achieved...
                                continue
                            else:
                                res_yearly_loss_values.append(interpolate_damages(rps, res_loss_curves[Pr_L], current_loss, protection_level))
                                com_yearly_loss_values.append(interpolate_damages(rps, com_loss_curves[Pr_L], current_loss, protection_level))
                                ind_yearly_loss_values.append(interpolate_damages(rps, ind_loss_curves[Pr_L], current_loss, protection_level))
                                inf_yearly_loss_values.append(interpolate_damages(rps, inf_loss_curves[Pr_L], current_loss, protection_level))
                    else:
                        # Independent simulation for this basin
                        r = random_numbers[simulation, year, i]
                        yearly_losses.append(r)
                        for Pr_L in res_loss_curves: # loop through basin protection levels
                            if Pr_L <= r:
                                continue
                            else:
                                res_yearly_loss_values.append(interpolate_damages(rps, res_loss_curves[Pr_L], r, protection_level))
                                com_yearly_loss_values.append(interpolate_damages(rps, com_loss_curves[Pr_L], r, protection_level))
                                ind_yearly_loss_values.append(interpolate_damages(rps, ind_loss_curves[Pr_L], r, protection_level))
                                inf_yearly_loss_values.append(interpolate_damages(rps, inf_loss_curves[Pr_L], r, protection_level))

            # Aggregate losses for the current year
            res_national_losses_per_year[simulation, year] = sum(res_yearly_loss_values)
            com_national_losses_per_year[simulation, year] = sum(com_yearly_loss_values)
            ind_national_losses_per_year[simulation, year] = sum(ind_yearly_loss_values)
            inf_national_losses_per_year[simulation, year] = sum(inf_yearly_loss_values)

    res_df = pd.DataFrame(res_national_losses_per_year, columns=[f'Year_{i+1}' for i in range(num_years)])
    com_df = pd.DataFrame(com_national_losses_per_year, columns=[f'Year_{i+1}' for i in range(num_years)])
    ind_df = pd.DataFrame(ind_national_losses_per_year, columns=[f'Year_{i+1}' for i in range(num_years)])
    inf_df = pd.DataFrame(inf_national_losses_per_year, columns=[f'Year_{i+1}' for i in range(num_years)])

    return res_df, com_df, ind_df, inf_df

# Necessary functions
def update_calibration_parameters(sheet, parameter, new_value):
    '''
    This function updates the calibration sheet in DIGNAD.
    '''

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == parameter:
                # Assuming the value needs to be updated in the cell right after the parameter
                target_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                try:
                    # Convert to float first
                    target_cell.value = float(new_value)
                except ValueError:
                    # if above doesn't work, save as string
                    target_cell.value = new_value
                return True  # Return after the first match to avoid unnecessary updates
    return False  # Return False if parameter not found

def update_natural_hazard_parameters(nat_disaster_year, recovery_period, tradable_impact,
                                    nontradable_impact, reconstruction_efficiency, public_debt_premium,
                                    public_impact, private_impact, share_tradable):
    '''
    This function returns a dictionary that will be used to populate the Disasters sheet in DIGNAD.
    It takes as input the 9 parameters the user typically has to set in the Disasters sheet.
    '''

    excel_updates = {
    (3, 4): nat_disaster_year - 2021, # C4 cell update
    (4, 4): nat_disaster_year, # D4 cell update
    (4, 7): tradable_impact, # D7 cell update
    (4, 8): nontradable_impact, # D8 cell update
    (4, 9): reconstruction_efficiency, # D9 cell update
    (4, 10): public_debt_premium, # D10 cell update
    (4, 11): public_impact, # D11 cell update
    (4, 12): private_impact, # D12 cell update
    (4, 13): share_tradable, # D13 cell update
    (3, 17): nat_disaster_year, # C17 cell update
    (4, 17): nat_disaster_year, # D17 cell update
    (3, 18): nat_disaster_year + recovery_period, # C18 cell update
    (4, 18): nat_disaster_year + recovery_period, # D18 cell update
    (3, 20): nat_disaster_year, # C20 cell update
    (3, 21): nat_disaster_year + recovery_period, # C21 cell update
    (3, 23): nat_disaster_year, # C23 cell update
    (3, 24): nat_disaster_year + recovery_period, # C24 cell update
    (3, 26): nat_disaster_year, # C26 cell update
    (3, 27): nat_disaster_year + recovery_period # C27 cell update
    }

    return excel_updates

def run_DIGNAD(calibration_csv, nat_disaster_year, recovery_period, tradable_impact, nontradable_impact,
                reconstruction_efficiency, public_debt_premium, public_impact, private_impact, share_tradable):
    '''
    This function runs on instance of DIGNAD with a prespecified calibration csv.
    Parameters passed to the function are the natural hazard parameters.
    Function outputs a list of GDP losses - from 2021 - 2040
    Need to set filepaths for own DIGNAD installation.
    '''

    ### 1. Load the original Excel file - this is where all DIGNAD parameters are set
    excel_file = r"D:\projects\sovereign-risk\Thailand\DIGNAD\DIGNAD_Toolkit_2023\PW_SHARED_2023\DIGNAD_Toolkit\DIGNAD_python\input_DIG-ND.xlsx"
    wb = load_workbook(excel_file)

    ### 2. Load the CSV with calibration parameters
    calibration_df = pd.read_csv(calibration_csv)

    ### 3. Set calibration parameters
    sheet = wb['Calibration']
    # Iterate over the calibration DataFrame rows
    for index, row in calibration_df.iterrows():
        parameter = row['Parameters']  # The column name in your CSV for the parameter names
        new_value = row['Values']       # The column name in your CSV for the new values
        updated = update_calibration_parameters(sheet, parameter, new_value)
        if not updated:
            print(f"Parameter '{parameter}' not found in the Excel sheet.")

    ### 4. Update disasters sheet
    natdisaster_params = update_natural_hazard_parameters(nat_disaster_year, recovery_period, tradable_impact,
                                                                nontradable_impact, reconstruction_efficiency,
                                                                public_debt_premium, public_impact, private_impact, share_tradable)
    sheet = wb['Disasters']
    for (col, row), value in natdisaster_params.items():
        cell = sheet.cell(row=row, column=col)
        cell.value = value

    ### 5. Save Excel Workbook
    wb.save(excel_file)

    ### 6. Run Matlab
    matlab_script = r"D:\projects\sovereign-risk\Thailand\DIGNAD\DIGNAD_Toolkit_2023\PW_SHARED_2023\DIGNAD_Toolkit\DIGNAD_python\simulate.m"
    result = subprocess.call(["matlab", "-batch", "run('" + matlab_script + "')"])
    if int(result) != 0:
        print("MATLAB script not executed succesfully")
        return None, None

    ### 7. Read results from Excel sheet
    # Get today's date as that is the name of file and directory
    today = datetime.datetime.today().strftime("%d%b%Y")
    file_path = r"D:\projects\sovereign-risk\Thailand\DIGNAD\DIGNAD_Toolkit_2023\PW_SHARED_2023\DIGNAD_Toolkit\DIGNAD_python\Excel output\%s\Model_output_%s.xlsx" % (today, today)
    df = pd.read_excel(file_path)
    years = list(df.iloc[:, 0])
    gdp_impact = list(df.iloc[:, 1])

    return gdp_impact, years
    